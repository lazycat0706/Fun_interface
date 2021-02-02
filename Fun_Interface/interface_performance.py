import json
import time

from gevent.threadpool import ThreadPoolExecutor

from Fun_Interface.common import get_env_url, get_app_header
import threading
# import grequests
# import sys
# sys.setrecursionlimit(10000)
import requests
import openpyxl

# personal_information
case_path = "D:/Python development/Fun_Interface/app_interface.xlsx"

# test_api = input("需要测试哪个接口：")
test_api = "personal_information"
env = "sit"
base_url = get_env_url(env)


class get_excel_data(object):
    def __init__(self, file):
        self.file = file
        self.wb = openpyxl.load_workbook(self.file)
        sheets = self.wb.sheetnames
        self.sheet = sheets[0]
        self.ws = self.wb[self.sheet]

    def get_interface_row_num(self, test_api):
        rows = self.ws.max_row
        for num in range(1, rows):
            api_name = self.ws["A" + str(num)].value
            if api_name == test_api:
                # print(api_name)
                return num

    def get_cell_value(self, row, column):
        cell_value = self.ws.cell(row=row, column=column).value
        return cell_value

    def get_row_value(self, row):
        columns = self.ws.max_column
        row_data = []
        for i in range(1, columns+1):
            cell_value = self.ws.cell(row=row, column=i).value
            row_data.append(cell_value)
        return row_data

    def get_col_value(self, column):
        rows = self.ws.max_row
        column_data = []
        for i in range(1, rows+1):
            cell_value = self.ws.cell(row=i, column=column).value
            column_data.append(cell_value)
        return column_data




# print(interface_info)
def get_response(interface_info):
    url = base_url + interface_info[2]
    headers = get_app_header()
    body_data = json.loads(interface_info[4])
    response = requests.get(url=url, headers=headers, params=body_data).text
    # print(response)
    # return response


def thread_pool_main(interface_info):
    thread_obj = ThreadPoolExecutor(max_workers=1, thread_name_prefix="WorkExecutor")
    # logger.info("Master ThreadPool Executor starts thread worker")
    thread_obj.submit(get_response(interface_info))











if __name__ == '__main__':
    excel = get_excel_data(case_path)
    num = excel.get_interface_row_num(test_api)
    interface_info = excel.get_row_value(num)
    process_num = 50
    threads = []
    time1 = time.perf_counter()
    for i in range(1, process_num+1):
        t = threading.Thread(target=get_response(interface_info), name="T"+str(i), daemon=True)
        print('thread %s' % t)
        # threads.append(t)
    # for t in threads:
    #     time.sleep(0.5)
    #     print('thread %s' % t)
    #     t.start()
    # t.join()

    time2 = time.perf_counter()
    times = time2-time1
    print(times)
