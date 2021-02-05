import json

import requests
import pandas as pd
import time
import pymysql
import openpyxl
from openpyxl.styles import PatternFill, Alignment
import smtplib
from email.mime.text import MIMEText
from email.utils import formataddr
from email.mime.multipart import MIMEMultipart
import traceback
import os


# 根据环境env值，返回接口固定URL
def get_env_url(env, qsxq_type):
    base_url = ""
    if qsxq_type == "app":
        if env == "sit":
            base_url = "https://sns-test.trendingstar.tech/"
            # return base_url
    elif qsxq_type == "erp":
        if env == "sit":
            base_url = "http://erp-server-test.unicornbpm.com/"
            # return base_url
    elif qsxq_type == "applet":
        if env == "sit":
            base_url = "https://wx-server-test.unicornbpm.com/"
            # return base_url
    return base_url


# 发送get请求
def send_get_request(url, headers, body_data):
    res = requests.get(url=url, headers=headers, params=body_data)
    return res.text


# 发送post请求，参数是params，提交在URL中
def send_post_params_request(url, headers, body_data):
    res = requests.post(url=url, headers=headers, params=body_data)
    return res.text


# 发送post请求，参数是json，以表单形式提交
def send_post_json_request(url, headers, body_data):
    res = requests.post(url=url, headers=headers, json=body_data)
    return res.text


# 获取登录url
def get_login_url(qsxq_type):
    login_url = ""
    if qsxq_type == "comm":
        login_url = "http://47.114.138.254:10380/api/v1/login"
    elif qsxq_type == "erp":
        login_url = "http://erp-server-test.unicornbpm.com/api/erp/login"
    return login_url


# 获取app请求头（iOS）
def get_headers(qsxq_type, token=""):
    # app请求头
    app_headers = {
        "accept-language": "zh-Hans-CN;q=1",
        "x-forwarded-for": "113.116.5.96",
        "version": "0.4.0",
        "accept": "*/*",
        "build": "1",
        "Content-type": "application/json",
        "connection": "close",
        "accept-encoding": "gzip, deflate, br",
        "app_client": "ios",
        "userid": "7356283a9c6405044fcdc6bec7421347",
        "appid": "f5cd51ef183ef0f5c93a79265a52a353",
        "user-agent": "qu shi xing qiu/0.4.0 (iPhone; iOS 14.2; Scale/3.00)",
    }
    # 小程序请求头
    applet_headers = {
        "userid": "57ef183771d9ac56905d706c0e185e0e",
        "appid": "wxd30c7522dd4574b8683e7fdc75d17a"
    }
    # 运营中心请求头
    comm_headers = {
        "Accept": "application/json, text/plain, */*",
        "Accept-Encoding": "gzip, deflate",
        "Accept-Language": "zh-CN, zh; q=0.9",
        "Authorization": token,
        "Connection": "keep-alive",
        "Content-Length": "381",
        "Content-Type": "application/json",
        "Host": "47.114.138.254:10380",
        "Origin": "http://47.114.138.254:10281",
        "Referer": "http://47.114.138.254:10281/",
        "User-Agent": "Mozilla/5.0(Windows NT 10.0; Win64; x64) AppleWebKit/537.36(KHTML, like Gecko) Chrome/88.0.4324.104 Safari/537.36"
    }
    # erp请求头
    erp_headers = {
        "accept": "application/json, text/plain, */*",
        "Accept - Encoding": "gzip, deflate",
        "Accept - Language": "zh-CN, zh; q=0.9",
        "Authorization": token,
        "Connection": "keep-alive",
        "Host": "erp-server-test.unicornbpm.com",
        "Origin": "http://47.115.5.180:10281",
        "Referer": "http://47.115.5.180:10281/",
        "User-Agent": "Mozilla / 5.0(Windows NT 10.0; Win64; x64) AppleWebKit/537.36(KHTML, like Gecko) Chrome / 87.0.4280.88 Safari / 537.36"
    }
    # 根据qsxq_type返回headers
    if qsxq_type == "app":
        return app_headers
    elif qsxq_type == "applet":
        return applet_headers
    elif qsxq_type == "comm":
        return comm_headers
    elif qsxq_type == "erp":
        return erp_headers


# 获取登录post_body数据
def get_login_data(login_key, qsxq_type):
    # 运营中心登录body
    comm_login_data = {
        "user_phone": "18611112222",
        "user_password": "12345678",
        "captcha": "captcha_code",
        "key": login_key,
    }
    # erp登录body
    erp_login_data = {
        "user_phone": "18688423735",

        "user_password": "12345678",
        "captcha": "captcha_code",
        "key": login_key,
    }
    # 判断qsxq_type返回登录body
    if qsxq_type == "comm":
        return comm_login_data
    if qsxq_type == "erp":
        return erp_login_data


# 获取ERP、运营中心的headers中的KEY，及返回的token
def get_login_token(qsxq_type):
    # 获取key的url
    key_url = "http://47.114.138.254:10380/api/v1/captcha"
    key_response = json.loads(requests.get(key_url).text)
    login_key = key_response["data"]["key"]
    login_url = get_login_url(qsxq_type)
    login_headers = get_headers(qsxq_type)
    login_data = get_login_data(login_key, qsxq_type)
    login_response = json.loads(requests.post(url=login_url, headers=login_headers, json=login_data).text)
    token = login_response.get("data").get("access_token")
    # 拼接token并返回
    token = "Bearer " + token
    print(token)
    return token


# 读取Excel数据中用例信息
def get_excel(case_path):
    temp_list = []
    data_frame = pd.read_excel(case_path)
    for data in data_frame.values:
        data = list(data)
        temp_list.append(data)
    return temp_list


# 获取各种时间
def get_time():
    timestamp = time.time()
    return timestamp


# 判断预期结果，并将错误的结果写入Excel中
def check(response, expect, loc_num, case_path):
    temp = []
    # 设置失败的用例背景色为红色
    fail_fill = PatternFill("solid", fgColor="FF0000")
    # 设置成功的用例背景色为蓝色
    pass_fill = PatternFill("solid", fgColor="1890FF")
    wb = openpyxl.load_workbook(case_path)
    ws = wb["Sheet1"]
    loc = 'g' + str(loc_num)
    fail_check = 0
    for k in expect:
        if (expect[k] == "" and response[k] is None) or (expect[k] == "" and response[k] == "null"):
            res = 'success'
            temp.append(res)
            print("接口正常，预期结果{%s}正确" % k)
        elif k in response and (expect[k] == response[k]):
            res = 'success'
            temp.append(res)
            print("接口正常，预期结果{%s}正确" % k)
        else:
            res = 'fail' + '--' + '预期结果{%s}的值不正确，expect:{%s}，response:{%s}' % (k, expect[k], response[k])
            temp.append(res)
            fail_check += 1
            print("预期结果不正确{%s}的值不正确，expect:{%s}，response:{%s}" % (k, expect[k], response[k]))
    # print(temp)
    if fail_check > 0:
        ws[loc].fill = fail_fill
    else:
        ws[loc].fill = pass_fill
    ws[loc] = str(temp)
    wb.save(case_path)


# 储存流程中需要的变量
def save_variable(api_name, response, loc_num, case_path):
    # 设置单元格对齐格式
    center_alignment = Alignment(horizontal='center', vertical='center')
    wb = openpyxl.load_workbook(case_path)
    ws = wb["Sheet2"]
    api_loc = 'a' + str(loc_num)
    response_loc = 'b' + str(loc_num)
    ws[api_loc] = api_name
    ws[api_loc].alignment = center_alignment
    ws[response_loc] = str(response)
    wb.save(case_path)


# 判断预期结果及保存过程中产生的返回信息
def check_save(loc_num, api_name, expect_result, response, case_path):
    check(response, expect_result, loc_num, case_path)
    save_variable(api_name, response, loc_num, case_path)
    print(response, expect_result)


# 连接数据库
def connect_db(sql, ex_type="query"):
    host = "112.124.90.5"
    username = "qsxq_rds_dev"
    password = "Qsxq_rds_dev"
    db = "qsxq_sns_kuplay_test"
    conn = pymysql.connect(host=host, user=username, password=password, db=db)
    cursor = conn.cursor()
    if ex_type == "query":
        cursor.execute(sql)
        while 1:
            result = cursor.fetchall()
            if result is None:
                break
            return result
    elif ex_type == "ex":
        ex_result = cursor.execute(sql)
        print("执行结果", ex_result)
        conn.commit()
        return ex_result
    conn.close()


# 发送邮件
# noinspection PyTypeChecker
def send_email(case_path):
    """发送邮件"""
    # 读取excel数据
    data = pd.read_excel(case_path, sheet_name="Sheet3", usecols=["收件人", "抄送人", "附件"], nrows=8).fillna(0)
    receiver = list(set(data.get("收件人").values))
    cc = list(set(data.get("抄送人").values))
    files = list(set(data.get("附件").values))
    if 0 in receiver:
        receiver.remove(0)
    if 0 in cc:
        cc.remove(0)
    if 0 in files:
        files.remove(0)
    # 基本信息
    my_sender = '542493741@qq.com'  # 发件人邮箱账号
    my_pass = 'uwvinkqvhwbibeeg'  # 发件人邮箱密码
    path = "D:/Python development/Fun_Interface"  # 附件路径

    message = MIMEMultipart()
    message['From'] = formataddr(["测试小组", my_sender])
    message['To'] = formataddr(["", ','.join(receiver)])
    message['Cc'] = formataddr(["", ','.join(cc)])
    message['Subject'] = "接口测试报告"

    # 正文内容
    msg = MIMEText("本次接口自动化运行结果已出，详情请查看附件", 'plain', 'utf-8')
    message.attach(msg)

    # 发送附件
    for file in files:
        if os.path.isfile(path + '/' + file):
            # 构造附件
            att = MIMEText(open(path + '/' + file, 'rb').read(), 'base64', 'utf-8')
            att["Content-Type"] = 'application/octet-stream'
            att.add_header("Content-Disposition", "attachment", filename=("utf-8", "", file))
            message.attach(att)

    # STMP服务
    try:
        server = smtplib.SMTP_SSL("smtp.qq.com", 465)
        server.login(my_sender, my_pass)
        server.sendmail(my_sender, receiver + cc, message.as_string())
        server.quit()
        return True
    except Exception as e:
        # traceback.print_exc()
        print(e)
        return False
